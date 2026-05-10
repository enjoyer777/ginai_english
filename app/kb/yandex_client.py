"""Загрузка xlsx-файла базы знаний с Яндекс.Диска через REST API.

Поток:
  1. GET https://cloud-api.yandex.net/v1/disk/resources/download?path=...
     с заголовком Authorization: OAuth <token>
     → возвращает href на временную ссылку для скачивания.
  2. GET по href → бинарные байты xlsx.
  3. openpyxl парсит листы → KBSnapshot.

OAuth-токен: см. https://yandex.ru/dev/disk/poligon/ или регистрация приложения
на https://oauth.yandex.ru/ с правом cloud_api:disk.read.
"""

from __future__ import annotations

import io
from datetime import datetime, time
from typing import Any

import httpx
from loguru import logger
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings
from app.kb.schema import (
    Course,
    Direction,
    FAQItem,
    KBSnapshot,
    Level,
    ScheduleSlot,
    Settings as KBSettings,
    Teacher,
)
from app.utils.working_hours import WEEKDAY_KEYS, parse_hours_range

YANDEX_API_BASE = "https://cloud-api.yandex.net/v1/disk"
HTTP_TIMEOUT = 30.0


class YandexDiskError(RuntimeError):
    pass


async def fetch_kb_xlsx() -> bytes:
    """Скачивает xlsx-файл целиком в память."""
    headers = {
        "Authorization": f"OAuth {settings.yandex_disk_oauth_token.get_secret_value()}",
        "Accept": "application/json",
    }
    params = {"path": settings.yandex_disk_file_path}

    async with httpx.AsyncClient(timeout=HTTP_TIMEOUT, follow_redirects=True) as client:
        # Шаг 1: получить временную ссылку
        r = await client.get(
            f"{YANDEX_API_BASE}/resources/download", headers=headers, params=params
        )
        if r.status_code != 200:
            raise YandexDiskError(
                f"Yandex.Disk download API error {r.status_code}: {r.text[:200]}"
            )
        href = r.json().get("href")
        if not href:
            raise YandexDiskError("Yandex.Disk did not return download href")

        # Шаг 2: скачать файл (без OAuth-заголовка — ссылка одноразовая, подписанная)
        r2 = await client.get(href)
        if r2.status_code != 200:
            raise YandexDiskError(f"Download failed {r2.status_code}: {r2.text[:200]}")
        return r2.content


async def load_kb_snapshot() -> KBSnapshot:
    raw = await fetch_kb_xlsx()
    return parse_xlsx(raw)


# --------- xlsx parser ---------

VALID_DIRECTIONS = {"работа", "переезд", "путешествия", "для себя"}
VALID_LEVELS = {"A0-A1", "A2", "B1", "B2"}


def parse_xlsx(blob: bytes) -> KBSnapshot:
    # Сначала пытаемся быстро (read_only=True — стримовое чтение, малая память).
    # Я.Документы и некоторые онлайн-редакторы при сохранении в xlsx иногда отдают
    # файл без styles.xml — read_only-режим openpyxl на этом падает с
    # 'could not read stylesheet from None'. На fallback идём в обычный режим,
    # который игнорирует такие огрехи.
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:
        logger.warning(
            "openpyxl read_only mode failed ({}); falling back to full-load mode",
            type(e).__name__,
        )
        wb = load_workbook(io.BytesIO(blob), read_only=False, data_only=True, keep_vba=False)

    courses = _parse_courses(wb["Курсы"]) if "Курсы" in wb.sheetnames else []
    schedules = _parse_schedules(wb["Расписание"]) if "Расписание" in wb.sheetnames else []
    teachers = _parse_teachers(wb["Преподаватели"]) if "Преподаватели" in wb.sheetnames else []
    faq = _parse_faq(wb["FAQ"]) if "FAQ" in wb.sheetnames else []
    settings_obj = _parse_settings(wb["Настройки"]) if "Настройки" in wb.sheetnames else KBSettings()

    if "Праздники" in wb.sheetnames:
        overrides, notes = _parse_holidays(wb["Праздники"])
        settings_obj.date_overrides = overrides
        settings_obj.date_notes = notes

    logger.info(
        "KB parsed: {} courses, {} schedule slots, {} teachers, {} faq, {} date overrides",
        len(courses),
        len(schedules),
        len(teachers),
        len(faq),
        len(settings_obj.date_overrides),
    )
    return KBSnapshot(
        courses=courses,
        schedules=schedules,
        teachers=teachers,
        faq=faq,
        settings=settings_obj,
    )


def _read_dict_rows(ws: Worksheet) -> list[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows, [])]
    out: list[dict[str, Any]] = []
    for row in rows:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return out


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return default


def _to_bool(v: Any) -> bool:
    s = _to_str(v).lower()
    return s in {"да", "yes", "true", "1", "y", "истина"}


def _to_date(v: Any):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month"):  # date
        return v
    s = _to_str(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_courses(ws: Worksheet) -> list[Course]:
    out: list[Course] = []
    for r in _read_dict_rows(ws):
        direction = _to_str(r.get("direction"))
        if direction not in VALID_DIRECTIONS:
            logger.warning("Skipping course with bad direction: {}", direction)
            continue
        levels_raw = _to_str(r.get("levels"))
        levels: list[Level] = []
        for lv in [x.strip() for x in levels_raw.split(",")]:
            if lv in VALID_LEVELS:
                levels.append(lv)  # type: ignore[arg-type]
        try:
            out.append(
                Course(
                    id=_to_str(r.get("id")),
                    name=_to_str(r.get("name")),
                    direction=direction,  # type: ignore[arg-type]
                    levels=levels,
                    duration_weeks=_to_int(r.get("duration_weeks")),
                    format=_to_str(r.get("format")),
                    price_rub=_to_int(r.get("price_rub")),
                    installment=_to_bool(r.get("installment")),
                    description=_to_str(r.get("description")),
                    nearest_start=_to_date(r.get("nearest_start")),
                )
            )
        except Exception as e:  # pragma: no cover
            logger.warning("Skipping malformed course row: {}", e)
    return out


def _parse_schedules(ws: Worksheet) -> list[ScheduleSlot]:
    out: list[ScheduleSlot] = []
    for r in _read_dict_rows(ws):
        start_date = _to_date(r.get("start_date"))
        if start_date is None:
            continue
        out.append(
            ScheduleSlot(
                course_id=_to_str(r.get("course_id")),
                start_date=start_date,
                days_time=_to_str(r.get("days_time")),
                teacher_name=_to_str(r.get("teacher_name")),
                seats_left=_to_int(r.get("seats_left")),
            )
        )
    return out


def _parse_teachers(ws: Worksheet) -> list[Teacher]:
    out: list[Teacher] = []
    for r in _read_dict_rows(ws):
        out.append(
            Teacher(
                name=_to_str(r.get("name")),
                experience_years=_to_int(r.get("experience_years")),
                specialization=_to_str(r.get("specialization")),
                bio=_to_str(r.get("bio")),
            )
        )
    return out


def _parse_faq(ws: Worksheet) -> list[FAQItem]:
    out: list[FAQItem] = []
    for r in _read_dict_rows(ws):
        q = _to_str(r.get("question"))
        a = _to_str(r.get("answer"))
        if q and a:
            out.append(FAQItem(question=q, answer=a))
    return out


def _parse_holidays(ws: Worksheet) -> tuple[
    dict[Any, tuple[time, time] | None],
    dict[Any, str],
]:
    """Лист 'Праздники' — список дат-исключений.

    Колонки:
      - date  (обязательно): YYYY-MM-DD или DD.MM.YYYY
      - hours (опционально): 'HH:MM-HH:MM' если в этот день работаем особым графиком,
                             пусто/отсутствует — нерабочий день
      - note  (опционально): человеческое описание (например, 'День Победы')
    """
    overrides: dict[Any, tuple[time, time] | None] = {}
    notes: dict[Any, str] = {}
    for r in _read_dict_rows(ws):
        d = _to_date(r.get("date"))
        if d is None:
            continue
        hours_raw = _to_str(r.get("hours"))
        overrides[d] = parse_hours_range(hours_raw) if hours_raw else None
        note = _to_str(r.get("note"))
        if note:
            notes[d] = note
    return overrides, notes


def _parse_settings(ws: Worksheet) -> KBSettings:
    """Лист 'Настройки' — две колонки key|value, плоский словарь.

    Распознаваемые ключи:
      - working_hours_{mon..sun} = "09:00-19:00" (пусто = выходной)
      - greeting_text, pii_disclaimer
      - contact_phone, contact_email, contact_site
      - social_telegram, social_vk, ...
    """
    rows = _read_dict_rows(ws)
    kv: dict[str, str] = {}
    for r in rows:
        key = _to_str(r.get("key"))
        value = _to_str(r.get("value"))
        if key:
            kv[key] = value

    working_hours: dict[str, tuple[time, time] | None] = {}
    for d in WEEKDAY_KEYS:
        working_hours[d] = parse_hours_range(kv.get(f"working_hours_{d}", ""))

    contacts = {
        k.removeprefix("contact_"): v for k, v in kv.items() if k.startswith("contact_")
    }
    socials = {
        k.removeprefix("social_"): v for k, v in kv.items() if k.startswith("social_")
    }

    return KBSettings(
        working_hours=working_hours,
        contacts=contacts,
        socials=socials,
        greeting_text=kv.get("greeting_text", ""),
        pii_disclaimer=kv.get("pii_disclaimer", ""),
    )
